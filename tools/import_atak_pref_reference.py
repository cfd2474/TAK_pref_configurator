#!/usr/bin/env python3
"""Import ATAK preference key reference from ATAK-Preferences-Key.xlsx."""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover - dev dependency
    raise SystemExit("Install openpyxl to run this tool: pip install openpyxl") from exc

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "backend" / "data" / "ATAK-Preferences-Key.xlsx"
OUTPUT = ROOT / "backend" / "data" / "atak_pref_reference.json"

JAVA_CLASSES = {
    "boolean": "class java.lang.Boolean",
    "integer": "class java.lang.Integer",
    "string": "class java.lang.String",
    "float": "class java.lang.Float",
    "select": "class java.lang.String",
}


def parse_bullets(raw: str) -> list[str]:
    if not raw:
        return []
    values = []
    for line in str(raw).splitlines():
        line = line.strip()
        if not line:
            continue
        if line.startswith("•"):
            line = line[1:].strip()
        values.append(line)
    return values


def parse_key_value_option(line: str) -> tuple[str, str] | None:
    if "=" not in line:
        return None
    value, label = line.split("=", 1)
    value = value.strip()
    label = label.strip()
    if not value or not label:
        return None
    return value, label


def normalize_reference_row(item: str, key: str, raw_type: str, raw_values: str) -> dict:
    exportable = True
    reference_type = raw_type.strip()
    bullets = parse_bullets(raw_values)
    options: list[dict[str, str]] = []
    reference_hint: str | None = None
    field_type = "string"
    storage_type = "string"
    input_mode: str | None = None

    if reference_type == "Boolean":
        field_type = "boolean"
        storage_type = "boolean"
        input_mode = "select"
        options = [{"label": "True", "value": "true"}, {"label": "False", "value": "false"}]
    elif reference_type == "Integer":
        field_type = "integer"
        storage_type = "integer"
        if bullets:
            reference_hint = bullets[0]
    elif reference_type == "Information":
        exportable = False
        field_type = "information"
    elif reference_type in {"Menu Item", "Category"}:
        exportable = False
        field_type = reference_type.lower().replace(" ", "_")
    elif reference_type == "Selection":
        exportable = False
        field_type = "selection_action"
    elif reference_type == "String":
        if len(bullets) == 1 and bullets[0] in {"Alphanumeric", "Numeric"}:
            reference_hint = bullets[0]
            field_type = "string"
            storage_type = "string"
        elif len(bullets) == 1 and re.search(r"\d", bullets[0]):
            reference_hint = bullets[0]
            field_type = "string"
            storage_type = "string"
        elif len(bullets) == 2 and {bullet.lower() for bullet in bullets} == {"true", "false"}:
            field_type = "boolean"
            storage_type = "boolean"
            input_mode = "select"
            options = [{"label": "True", "value": "true"}, {"label": "False", "value": "false"}]
        elif len(bullets) > 1:
            field_type = "select"
            storage_type = "string"
            input_mode = "select"
            for bullet in bullets:
                parsed = parse_key_value_option(bullet)
                if parsed:
                    value, label = parsed
                    options.append({"label": label, "value": value})
                else:
                    options.append({"label": bullet, "value": bullet})
        else:
            field_type = "string"
            storage_type = "string"
    else:
        reference_hint = raw_values or None

    entry = {
        "item": item.strip(),
        "key": key.strip(),
        "reference_type": reference_type,
        "type": field_type,
        "exportable": exportable,
        "reference_hint": reference_hint,
    }
    if options:
        entry["options"] = options
        entry["input"] = input_mode or "select"
    if storage_type:
        entry["storage_type"] = storage_type
        java_class = JAVA_CLASSES.get(storage_type)
        if java_class:
            entry["java_class"] = java_class
    return entry


def load_workbook_rows(path: Path) -> list[tuple]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["ATAK Core"]
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    return rows


def build_reference(path: Path) -> dict:
    rows = load_workbook_rows(path)
    keys: dict[str, dict] = {}
    categories: dict[str, dict] = {}
    current_category_key: str | None = None
    current_category_title: str | None = None
    current_menu_key: str | None = None
    current_menu_title: str | None = None

    for row in rows[1:]:
        if not row or not row[1]:
            continue
        item = str(row[0] or "").strip()
        key = str(row[1]).strip()
        raw_type = str(row[2] or "").strip()
        raw_values = str(row[3] or "").strip()

        if raw_type == "Category":
            current_category_key = key
            current_category_title = item or key
            categories[key] = {
                "title": current_category_title,
                "key": key,
                "menu_key": current_menu_key,
                "menu_title": current_menu_title,
            }
            continue

        if raw_type == "Menu Item":
            current_menu_key = key
            current_menu_title = item or key
            continue

        entry = normalize_reference_row(item, key, raw_type, raw_values)
        entry["category_key"] = current_category_key
        entry["category_title"] = current_category_title
        entry["menu_key"] = current_menu_key
        entry["menu_title"] = current_menu_title
        keys[key] = entry

    exportable_count = sum(1 for entry in keys.values() if entry.get("exportable"))
    return {
        "source": path.name,
        "sheet": "ATAK Core",
        "keys": keys,
        "categories": categories,
        "stats": {
            "total_keys": len(keys),
            "exportable_keys": exportable_count,
            "category_count": len(categories),
        },
    }


def main() -> int:
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx_path.exists():
        raise SystemExit(f"Reference workbook not found: {xlsx_path}")

    reference = build_reference(xlsx_path)
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(reference, indent=2), encoding="utf-8")
    print(f"Wrote {OUTPUT}")
    print(json.dumps(reference["stats"], indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
